import openpyxl
import datetime


def readExcel():
    wb = openpyxl.load_workbook('/Users/wangpengju/PycharmProjects/Mnt_Inf_Platform/Data/data1.xlsx')
    print(type(wb))
    # sheets = wb.get_sheet_by_name('销售绩效清单')
    sheets = wb['Sheet1']
    a = 1
    b = 1
    accoutapplydata = []
    while b < 90:
        cell = sheets.cell(row=b, column=a)
        cell = cell.value
        b = b + 1
        accoutapplydata += [cell]
    print(accoutapplydata)
    return accoutapplydata


def writeExcel(i, ain, bin, cin, din, ein):
    wb = openpyxl.load_workbook('/Users/wangpengju/PycharmProjects/Mnt_Inf_Platform/Data/data3.xlsx')
    sheet = wb['Sheet1']
    sheet.append([i, ain, bin, cin, din, ein])
    wb.save('/Users/wangpengju/PycharmProjects/Mnt_Inf_Platform/Data/data3.xlsx')


if __name__ == "__main__":
    writeExcel('http://www.chinabidding.cc/read-329119463.html', '玉屏侗族自治县人民法院关于杨某名下位于现代城的房产第一次拍卖（第一次拍卖）的公告', '[贵州]', '2022-08-17')
